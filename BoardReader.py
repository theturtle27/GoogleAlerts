from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
import numpy as np
from datetime import datetime
from selenium.webdriver.common.keys import Keys
import openpyxl
# TODO:
#   detect english only websites
#   check for duplicates in google alerts

# path where chromedriver is located
PATH = r"chromedriver"

# keyword for searching
wb_obj = openpyxl.load_workbook("Parameters.xlsx")

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
keyword = sheet_obj.cell(row=2, column=5).value
numberOfArticles = int(sheet_obj.cell(row=3, column=5).value)

now = datetime.now()
date = now.strftime("%d-%m-%y %H%M")
# open chrome
browser = webdriver.Chrome(PATH)
# go to website
browser.get('https://boardreader.com/')

# locate the search box and send keyword
element = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[1]/div/div/div/div[2]/div/div[2]/input"))
)
element.send_keys(keyword)
# create dataFrame
dfMain = pd.DataFrame()
dfCopy = pd.DataFrame()
# empty list to store info
titlesText = []
sourcesText = []
linksText = []
keywordText = []
paragraphText = []
# i is variable for articles in current page
i = 0
j = 0
k = 0
WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "title")))

# get results and sources in a list
titles = browser.find_elements_by_class_name("title")
sources = browser.find_elements_by_class_name("last-info")
links = browser.find_elements_by_tag_name("a")

cookies = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/table/tbody/tr/td[2]/a")))
cookies.click()

while len(dfMain.index) < numberOfArticles:
    paragraph_summary = ""
    try:
        titlesText.append(titles[i].text)
        sourcesText.append(sources[i].text)
        linksText.append(links[(i*3)+1].get_attribute("href"))
        keywordText.append(keyword)
    except Exception as e:
        print(e)
        break
    # add paragraph
    try:
        links[(i*3)+1].send_keys(Keys.CONTROL + Keys.RETURN)
        # Switch to the new window and open URL B
        browser.switch_to.window(browser.window_handles[1])
        WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, "p")))
        paragraphs = browser.find_elements_by_tag_name("p")
        for paragraph in paragraphs:
            paragraph_summary = paragraph_summary + paragraph.text
    except Exception as e:
        print(e)
    finally:
        paragraphText.append(paragraph_summary)
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        i += 1

    # next page
    if i == 10:
        if k == 0:
            nextPage = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[1]/app-root/results/div/div[1]/div/div/div/p[2]/a")))
            k += 1
        else:
            nextPage = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[1]/app-root/results/div/div[1]/div/div/div/p[2]/a[2]")))
        nextPage.click()
        WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "title")))
        titles = browser.find_elements_by_class_name("title")
        sources = browser.find_elements_by_class_name("last-info")
        links = browser.find_elements_by_tag_name("a")
        i = 0
    dfCopy['Keyword'] = keywordText
    dfCopy['Title'] = titlesText
    dfCopy['Source'] = sourcesText
    dfCopy['Link'] = linksText
    dfCopy['Paragraph'] = paragraphText
    # drop duplicates
    dfMain = dfMain.append(dfCopy)
    dfMain.drop_duplicates(subset='Paragraph', inplace=True)
    print(len(dfMain.index), ' article done!')
    dfCopy = pd.DataFrame()
    keywordText = []
    titlesText = []
    sourcesText = []
    linksText = []
    paragraphText = []

try:
    wb = openpyxl.load_workbook(keyword + '.xlsx')

    # Inside this context manager, handle everything related to writing new data to the file\
    # without overwriting existing data
    with pd.ExcelWriter(keyword + '.xlsx', engine='openpyxl') as writer:
        # Your loaded workbook is set as the "base of work"
        writer.book = wb

        # Loop through the existing worksheets in the workbook and map each title to\
        # the corresponding worksheet (that is, a dictionary where the keys are the\
        # existing worksheets' names and the  values are the actual worksheets)
        writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}

        # Write the new data to the file without overwriting what already exists
        dfMain.to_excel(writer, date, index=False)

        # Save the file
        writer.save()
except Exception as e:
    dfMain.to_excel(keyword + '.xlsx',
                sheet_name=date,
                index=False)
dfMain.to_csv(keyword + date + '.csv', index=False)
