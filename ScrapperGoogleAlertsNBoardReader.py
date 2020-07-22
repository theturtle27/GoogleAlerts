from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
from datetime import datetime
import pandas as pd

# path where chromedriver is located
PATH = r"chromedriver"
# keyword for searching
wb_obj = openpyxl.load_workbook("Parameters.xlsx")

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
keyword = sheet_obj.cell(row=2, column=2).value
numberOfArticles = int(sheet_obj.cell(row=3, column=2).value)
questionNumber1 = sheet_obj.cell(row=1, column=8).value
questionNumber2 = sheet_obj.cell(row=2, column=8).value
questionNumber3 = sheet_obj.cell(row=3, column=8).value
questionNumber4 = sheet_obj.cell(row=4, column=8).value

# get date time
now = datetime.now()
date = now.strftime("%d-%m-%y %H%M")

# open chrome
browser = webdriver.Chrome(PATH)

# go to website
browser.get('https://www.google.com.sg/alerts#')

# locate the search box and send keyword
element = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "//*[@id='query_div']/input"))
)
element.send_keys(keyword)

# change results to all results
show_options = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located(
        (By.XPATH, "//*[@id='create-alert-div']/div[2]/div[2]/div[2]/span[3]")))
show_options.click()

show = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "//*[@id=':6']"))
)
show.click()

show_expanded = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "//*[@id=':5']"))
)
show_expanded.click()

# wait for page to load
WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "result_title_link")))

# get results and sources
links = browser.find_elements_by_class_name("result_title_link")
sources = browser.find_elements_by_class_name("result_source")

# empty list to store info
titlesText = []
sourcesText = []
linksText = []
keywordText = []
paragraphText = []

# variable to iterate through
i = 0
while i < numberOfArticles:
    paragraph_summary = ""
    try:
        titlesText.append(links[i].text)
        sourcesText.append(sources[i].text)
        linksText.append(links[i].get_attribute("href"))
        keywordText.append(keyword)
    except Exception as e:
        print(e)
        break
    # add paragraph
    try:
        links[i].send_keys(Keys.CONTROL + Keys.RETURN)
        # Switch to the new window and open URL B
        browser.switch_to.window(browser.window_handles[1])
        WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.TAG_NAME, "p")))
        paragraphs = browser.find_elements_by_tag_name("p")
        for paragraph in paragraphs:
            paragraph_summary = paragraph_summary + paragraph.text
        print(i + 1, ' article done!')
    except Exception as e:
        print(e)
        print(i + 1, ' article error!')
    finally:
        paragraphText.append(paragraph_summary)
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        i += 1

df = pd.DataFrame()
df['Keyword'] = keywordText
df['Title'] = titlesText
df['Source'] = sourcesText
df['Link'] = linksText
df['Paragraph'] = paragraphText

browser.get('https://boardreader.com/')

# locate the search box and send keyword
element = WebDriverWait(browser, 5).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[1]/div/div/div/div[2]/div/div[2]/input"))
)
element.send_keys(keyword)
# create dataFrame
dfCopy = pd.DataFrame()
# empty list to store info
titlesText = []
sourcesText = []
linksText = []
keywordText = []
paragraphText = []
# i is variable for articles in current page
i = 0
j = 0
k = 0
WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "title")))

# get results and sources in a list
titles = browser.find_elements_by_class_name("title")
sources = browser.find_elements_by_class_name("last-info")
links = browser.find_elements_by_tag_name("a")

cookies = WebDriverWait(browser, 5).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/table/tbody/tr/td[2]/a")))
cookies.click()

numAlertsArticle = len(df.index)

while (len(df.index) - numAlertsArticle) < numberOfArticles:
    paragraph_summary = ""
    try:
        titlesText.append(titles[i].text)
        sourcesText.append(sources[i].text)
        linksText.append(links[(i * 3) + 1].get_attribute("href"))
        keywordText.append(keyword)
    except Exception as e:
        print(e)
        break
    # add paragraph
    try:
        links[(i * 3) + 1].send_keys(Keys.CONTROL + Keys.RETURN)
        # Switch to the new window and open URL B
        browser.switch_to.window(browser.window_handles[1])
        WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.TAG_NAME, "p")))
        paragraphs = browser.find_elements_by_tag_name("p")
        for paragraph in paragraphs:
            paragraph_summary = paragraph_summary + paragraph.text
    except Exception as e:
        print(e)
    finally:
        paragraphText.append(paragraph_summary)
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        i += 1

    # next page
    try:
        if i == 10:
            if k == 0:
                nextPage = WebDriverWait(browser, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "/html/body/div[1]/app-root/results/div/div[1]/div/div/div/p[2]/a")))
                k += 1
            else:
                if k < 7:
                    nextPage = WebDriverWait(browser, 5).until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "#form-holder > div > div > p.pager > a:nth-child(12)")))
                else:
                    nextPage = WebDriverWait(browser, 5).until(
                        EC.presence_of_element_located(
                            (By.XPATH, "// *[ @ id = 'form-holder'] / div / div / p[2] / a[3]")))
            nextPage.click()
            WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.CLASS_NAME, "title")))
            titles = browser.find_elements_by_class_name("title")
            sources = browser.find_elements_by_class_name("last-info")
            links = browser.find_elements_by_tag_name("a")
            k += 1
            i = 0
    except Exception as e:
        print(e)
    dfCopy['Keyword'] = keywordText
    dfCopy['Title'] = titlesText
    dfCopy['Source'] = sourcesText
    dfCopy['Link'] = linksText
    dfCopy['Paragraph'] = paragraphText
    # drop duplicates
    df = df.append(dfCopy)
    df.drop_duplicates(subset='Paragraph', inplace=True)
    print(len(df.index), ' article done!')
    dfCopy = pd.DataFrame()
    keywordText = []
    titlesText = []
    sourcesText = []
    linksText = []
    paragraphText = []

browser.quit()
# drop duplicates
df.drop_duplicates(subset='Paragraph', inplace=True)
# look for existing file
excelPath = r"Excel\\"
try:
    wb = openpyxl.load_workbook(excelPath + keyword + '.xlsx')

    # Inside this context manager, handle everything related to writing new data to the file\
    # without overwriting existing data
    with pd.ExcelWriter(excelPath + keyword + '.xlsx', engine='openpyxl') as writer:
        # Your loaded workbook is set as the "base of work"
        writer.book = wb

        # Loop through the existing worksheets in the workbook and map each title to\
        # the corresponding worksheet (that is, a dictionary where the keys are the\
        # existing worksheets' names and the values are the actual worksheets)
        writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}

        # Write the new data to the file without overwriting what already exists
        df.to_excel(writer, date, index=False)

        # Save the file
        writer.save()
# write new file
except Exception as e:
    print(e)
    df.to_excel(excelPath + keyword + '.xlsx',
                sheet_name=date,
                index=False)

# write to csv
csvPath = "CSV\\"
df.to_csv(csvPath + keyword + date + '.csv', index=False)