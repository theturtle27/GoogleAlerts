# ---------------------------------------------------------------------------------------
# clear items from here
# ---------------------------------------------------------------------------------------
# C:\Users\NYC\Downloads
# D:\Work\YCS Survey Automation\topic_model\output

from datetime import datetime
import pandas as pd
import os
import gensim
import gensim.corpora as corpora
from gensim.utils import simple_preprocess
from gensim.models import CoherenceModel
import openpyxl
import nltk

nltk.download("popular")
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)
import pyLDAvis
import pyLDAvis.gensim
from nltk.corpus import stopwords

stop_words = stopwords.words('english')
stop_words.extend(
    ['me', 'and', 'or', 'i', 'I', 'you', 'u', 'we', 'We', 'us', 'he', 'she', 'it', 'they', 'them', 'would', 'got',
     'went', 'could', 'made', 'give', 'go', 'get', 'also', 'much', 'must', 'well', 'much', 'nil', 'ok'])
from tabulate import tabulate


# Defining functions
def format_topics_sentences(ldamodel=None, corpus=None, texts=None, sourcedata=None):
    # Init output
    sent_topics_df = pd.DataFrame()

    # Get main topic in each document
    for i, row_list in enumerate(ldamodel[corpus]):
        row = row_list[0] if ldamodel.per_word_topics else row_list
        # print(row)
        row = sorted(row, key=lambda x: (x[1]), reverse=True)
        # Get the Dominant topic, Perc Contribution and Keywords for each document
        for j, (topic_num, prop_topic) in enumerate(row):
            if j == 0:  # => dominant topic
                wp = ldamodel.show_topic(topic_num)
                topic_keywords = ", ".join([word for word, prop in wp])
                sent_topics_df = sent_topics_df.append(
                    pd.Series([int(topic_num), round(prop_topic, 4), topic_keywords]), ignore_index=True)
            else:
                break
    sent_topics_df.columns = ['Dominant_Topic', 'Perc_Contribution', 'Topic_Keywords']

    # Add original text to the end of the output
    contents = pd.Series(sourcedata)
    sent_topics_df = pd.concat([sent_topics_df, contents], axis=1)
    return (sent_topics_df)


def my_preprocess(text):
    return simple_preprocess(text, deacc=True)  # deacc=True removes punctuations


def word_swap(text):
    # for item in text.split():
    if len(text.split()) < 5:
        text = " "
    else:
        text = text
    return (text)


'''
datetoday = input("Please enter today's date in 6 digits:\n")
print("")
filename = input("Please enter the name of the file without the extension:\n")
print("")
numtopics = input("Please enter the number of word clusters to generate:\n")
print("")
col_name = input("Please enter column name:\n")
'''

datetoday = datetime.today().strftime('%d%m%y')

# open parameters workbook
wb_obj = openpyxl.load_workbook("Parameters.xlsx")

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# get filename, sheetName, number of topics, number of questions and question header from sheet
filename = sheet_obj.cell(row=14, column=2).value
sheetName = sheet_obj.cell(row=15, column=2).value
numtopics = int(sheet_obj.cell(row=16, column=2).value)
noOfQuestions = int(sheet_obj.cell(row=17, column=2).value)
header1 = sheet_obj.cell(row=18, column=2).value
header2 = sheet_obj.cell(row=19, column=2).value
header3 = sheet_obj.cell(row=20, column=2).value
header4 = sheet_obj.cell(row=21, column=2).value

OUTPUT_DIR = "output/"
current_directory = os.getcwd()
# directory = os.chdir(current_directory + "/rawdata")
directory = os.chdir("Excel")
data = pd.read_excel(filename + ".xlsx", sheetName)

'''
"C01: 1.	What was the most challenging moment during the course? How did you overcome it?"	
"C02: 2.	How was your experience of going through the course with peers from a different school? What went well or did not go well?"	
"C03: 3.	What have you learnt about the environment / community?"	
"C04: 4.	What personal goals / resolutions do you have after leaving this course?"	
"D02.reason: 2.	If you could choose from the start, would you still choose to go through the MOE-OBS Challenge Programme?"	
"D03.reason: 3.	Do you think the MOE-OBS Challenge Course should be a shared experience for all Secondary 3 youths? "	
"F03: 3.	Do you have any other comments or suggestions for improvement? Be as specific as you can!"
'''

# columns_to_LDA = data[[col_name]]
# columns_to_LDA = data[['Q69_11','Q69_12','Q70_12','Q70_13','Q93_29','Q93_35','Q94_30','Q94_36']]
# columns_to_LDA = data[['(Q69) 12. Do you have suggestions for improvement for any of the Mission X activities?']]
if noOfQuestions == 1:
    columns_to_LDA = data[[header1]]
elif noOfQuestions == 2:
    columns_to_LDA = data[[header1], [header2]]
elif noOfQuestions == 3:
    columns_to_LDA = data[[header1], [header2], [header3]]
elif noOfQuestions == 4:
    columns_to_LDA = data[[header1], [header2], [header3], [header4]]
# (Q69) 11. Do you have suggestions for improvement for any of the Mission X activities?    - BR1 - Q69_11
# (Q69) 12. Do you have suggestions for improvement for any of the Mission X activities?    - FV1 - Q69_12
# (Q70) 12. What other topics would you like to be included in the training workshop?       - BS1 - Q70_12
# (Q70) 13. What other topics would you like to be included in the training workshop?       - FW1 - Q70_13
# (Q93) 29. What were your key takeaways?                                                   - DA1 - Q93_29
# (Q93) 35. What were your key takeaways?                                                   - HJ1 - Q93_35
# (Q94) 30. Any other comments or areas of improvements for Mission X?                      - DB1 - Q94_30
# (Q94) 36. Any other comments or areas of improvements for Mission X?                      - HK1 - Q94_36

os.chdir(current_directory)

list_of_topictables = []
list_of_df_dominant_topic = []
df = columns_to_LDA

# WORD THRESHOLD - Only analyse sentences which have more than 5 words
for column in columns_to_LDA.columns:
    df['source'] = columns_to_LDA[column].fillna('').astype(str).apply(word_swap)
    df['clean'] = df['source'].apply(my_preprocess)
    bigram = gensim.models.Phrases(df['clean'], min_count=2, threshold=10)  # higher threshold fewer phrases.
    # Faster way to get a sentence clubbed as a trigram/bigram
    bigram_mod = gensim.models.phrases.Phraser(bigram)


    # STEP 4 REMOVE STOPWORDS, MAKE BIGRAMS AND LEMMATIZE -----------------------------
    def remove_stopwords(list_of_words):
        return [word for word in list_of_words if word not in stop_words]


    df['clean'] = df['clean'].apply(remove_stopwords)
    df['clean'] = df['clean'].apply(lambda list_of_words: bigram_mod[list_of_words])

    # STEP 11 CREATE THE DICTIONARY AND CORPUS NEEDED FOR TOPIC MODELLING -----------------------------
    # Create Dictionary
    id2word = corpora.Dictionary(df['clean'])

    # Create Corpus
    texts = df['clean']

    # Term Document Frequency
    corpus = [id2word.doc2bow(text) for text in texts]

    # STEP 5 BUILDING THE TOPIC MODEL -----------------------------
    # Build LDA model
    lda_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=numtopics, random_state=100,
                                                update_every=1, chunksize=100, passes=50, alpha='auto',
                                                per_word_topics=True)
    # cm = CoherenceModel(model=lda_model, corpus=corpus, texts=texts, coherence='c_v', dictionary=id2word)
    # print(cm)

    top_words_per_topic = []

    for t in range(lda_model.num_topics):
        top_words_per_topic.extend([(t,) + x for x in lda_model.show_topic(t, topn=10)])
    topictable = pd.DataFrame(top_words_per_topic, columns=['Topic', 'Word', 'Weight'])
    topictable.columns = [column + '_' + str(col_header) for col_header in topictable.columns]
    list_of_topictables.append(topictable)

    # Visualize the topics using pyLDAvis
    vis = pyLDAvis.gensim.prepare(lda_model, corpus, id2word)

    # Save the visualisation into an interactive page. Details on how to read it here: https://www.machinelearningplus.com/nlp/topic-modeling-gensim-python/
    pyLDAvis.save_html(vis, f'{OUTPUT_DIR}{filename}_{column}_LDA_vis_{datetoday}.html')

    df_topic_sents_keywords = format_topics_sentences(ldamodel=lda_model,
                                                      corpus=corpus,
                                                      texts=df['clean'],
                                                      sourcedata=df['source'])  # changed from df[column]

    # Format
    df_dominant_topic = df_topic_sents_keywords.reset_index()
    df_dominant_topic.columns = ['Document_No', 'Dominant_Topic', 'Topic_Perc_Contrib', 'Keywords', 'Text']
    df_dominant_topic.columns = [column + '_' + str(col_header) for col_header in df_dominant_topic.columns]
    list_of_df_dominant_topic.append(df_dominant_topic)

# write the topic tables and topic master matrix into one dataframe
alltopicstables = pd.DataFrame()
alltopictables = pd.concat(list_of_topictables, axis=1)
mastermatrix = pd.DataFrame()
mastermatrix = pd.concat(list_of_df_dominant_topic, axis=1)
outputmatrix = pd.concat([data, mastermatrix], axis=1)

# Writing the dataframes into a single Excel workbook ------------------------
outputfilename = f'{OUTPUT_DIR}{filename}{datetoday}_export.xlsx'
outputfilename_wo_ext = f'{filename}{datetoday}_export'
writer = pd.ExcelWriter(outputfilename, engine='xlsxwriter')
outputmatrix.to_excel(writer, sheet_name="Masterlist")
mastermatrix.to_excel(writer, sheet_name="Topic Master Matrix")
alltopictables.to_excel(writer, sheet_name="Topic Tables")
writer.save()
print(f'All done')

# ------------------------------------------------------------------------
# call upon the new viz script
# python topic_model_create_charts.py "What were your key takeaways?" "filename160320_export"
# https://raspberrypi.stackexchange.com/questions/17017/how-do-i-run-a-command-line-command-in-a-python-script
# ------------------------------------------------------------------------
# os.system("python topic_model_create_charts.py \"What were your key takeaways?\" \"filename090320_export\"")
# os.system(' python topic_model_create_charts.py "What were your key takeaways?" "filename090320_export" ')
# os.system(' python topic_model_create_charts.py "What were your key takeaways?" "' + outputfilename_
