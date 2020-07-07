from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
from selenium.webdriver.common.keys import Keys
import openpyxl
from datetime import datetime
import re
import numpy as np
import pandas as pd
import csv
import os
import gensim
import gensim.corpora as corpora
from gensim.utils import simple_preprocess
from gensim.models import CoherenceModel
import nltk
nltk.download("popular")
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import pyLDAvis
import pyLDAvis.gensim
from nltk.corpus import stopwords

# path where chromedriver is located
PATH = r"chromedriver"
# keyword for searching
wb_obj = openpyxl.load_workbook("Parameters.xlsx")

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
keyword = sheet_obj.cell(row=2, column=2).value
numberOfArticles = int(sheet_obj.cell(row=3, column=2).value)
questionNumber1 = sheet_obj.cell(row=1, column=8).value
questionNumber2 = sheet_obj.cell(row=2, column=8).value
questionNumber3 = sheet_obj.cell(row=3, column=8).value
questionNumber4 = sheet_obj.cell(row=4, column=8).value

# get date time
now = datetime.now()
date = now.strftime("%d-%m-%y %H%M")

# open chrome
browser = webdriver.Chrome(PATH)

# go to website
browser.get('https://www.google.com.sg/alerts#')

# locate the search box and send keyword
element = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "//*[@id='query_div']/input"))
)
element.send_keys(keyword)

# change results to all results
show_options = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located(
        (By.XPATH, "//*[@id='create-alert-div']/div[2]/div[2]/div[2]/span[3]")))
show_options.click()

show = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "//*[@id=':6']"))
)
show.click()

show_expanded = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "//*[@id=':5']"))
)
show_expanded.click()

# wait for page to load
WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "result_title_link")))

# get results and sources
links = browser.find_elements_by_class_name("result_title_link")
sources = browser.find_elements_by_class_name("result_source")

# empty list to store info
titlesText = []
sourcesText = []
linksText = []
keywordText = []
paragraphText = []

# variable to iterate through
i = 0
while i < numberOfArticles:
    paragraph_summary = ""
    try:
        titlesText.append(links[i].text)
        sourcesText.append(sources[i].text)
        linksText.append(links[i].get_attribute("href"))
        keywordText.append(keyword)
    except Exception as e:
        print(e)
        break
    # add paragraph
    try:
        links[i].send_keys(Keys.CONTROL + Keys.RETURN)
        # Switch to the new window and open URL B
        browser.switch_to.window(browser.window_handles[1])
        WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, "p")))
        paragraphs = browser.find_elements_by_tag_name("p")
        for paragraph in paragraphs:
            paragraph_summary = paragraph_summary + paragraph.text
        print(i + 1, ' article done!')
    except Exception as e:
        print(e)
        print(i + 1, ' article error!')
    finally:
        paragraphText.append(paragraph_summary)
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        i += 1

df = pd.DataFrame()
df['Keyword'] = keywordText
df['Title'] = titlesText
df['Source'] = sourcesText
df['Link'] = linksText
df['Paragraph'] = paragraphText

browser.get('https://boardreader.com/')

# locate the search box and send keyword
element = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[1]/div[1]/div/div/div/div[2]/div/div[2]/input"))
)
element.send_keys(keyword)
# create dataFrame
dfCopy = pd.DataFrame()
# empty list to store info
titlesText = []
sourcesText = []
linksText = []
keywordText = []
paragraphText = []
# i is variable for articles in current page
i = 0
j = 0
k = 0
WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "title")))

# get results and sources in a list
titles = browser.find_elements_by_class_name("title")
sources = browser.find_elements_by_class_name("last-info")
links = browser.find_elements_by_tag_name("a")

cookies = WebDriverWait(browser, 10).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/table/tbody/tr/td[2]/a")))
cookies.click()

numAlertsArticle = len(df.index)

while (len(df.index) - numAlertsArticle) < numberOfArticles:
    paragraph_summary = ""
    try:
        titlesText.append(titles[i].text)
        sourcesText.append(sources[i].text)
        linksText.append(links[(i * 3) + 1].get_attribute("href"))
        keywordText.append(keyword)
    except Exception as e:
        print(e)
        break
    # add paragraph
    try:
        links[(i * 3) + 1].send_keys(Keys.CONTROL + Keys.RETURN)
        # Switch to the new window and open URL B
        browser.switch_to.window(browser.window_handles[1])
        WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.TAG_NAME, "p")))
        paragraphs = browser.find_elements_by_tag_name("p")
        for paragraph in paragraphs:
            paragraph_summary = paragraph_summary + paragraph.text
    except Exception as e:
        print(e)
    finally:
        paragraphText.append(paragraph_summary)
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        i += 1

    # next page
    if i == 10:
        if k == 0:
            nextPage = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, "/html/body/div[1]/app-root/results/div/div[1]/div/div/div/p[2]/a")))
            k += 1
        else:
            if k < 7:
                nextPage = WebDriverWait(browser, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "#form-holder > div > div > p.pager > a:nth-child(12)")))
            else:
                nextPage = WebDriverWait(browser, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "// *[ @ id = 'form-holder'] / div / div / p[2] / a[3]")))
        nextPage.click()
        WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "title")))
        titles = browser.find_elements_by_class_name("title")
        sources = browser.find_elements_by_class_name("last-info")
        links = browser.find_elements_by_tag_name("a")
        k += 1
        i = 0
    dfCopy['Keyword'] = keywordText
    dfCopy['Title'] = titlesText
    dfCopy['Source'] = sourcesText
    dfCopy['Link'] = linksText
    dfCopy['Paragraph'] = paragraphText
    # drop duplicates
    df = df.append(dfCopy)
    df.drop_duplicates(subset='Paragraph', inplace=True)
    print(len(df.index), ' article done!')
    dfCopy = pd.DataFrame()
    keywordText = []
    titlesText = []
    sourcesText = []
    linksText = []
    paragraphText = []

browser.quit()
# drop duplicates
df.drop_duplicates(subset='Paragraph', inplace=True)
# look for existing file
excelPath = r"Excel\\"
try:
    wb = openpyxl.load_workbook(excelPath + keyword + '.xlsx')

    # Inside this context manager, handle everything related to writing new data to the file\
    # without overwriting existing data
    with pd.ExcelWriter(excelPath + keyword + '.xlsx', engine='openpyxl') as writer:
        # Your loaded workbook is set as the "base of work"
        writer.book = wb

        # Loop through the existing worksheets in the workbook and map each title to\
        # the corresponding worksheet (that is, a dictionary where the keys are the\
        # existing worksheets' names and the values are the actual worksheets)
        writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}

        # Write the new data to the file without overwriting what already exists
        df.to_excel(writer, date, index=False)

        # Save the file
        writer.save()
# write new file
except Exception as e:
    print(e)
    df.to_excel(excelPath + keyword + '.xlsx',
                sheet_name=date,
                index=False)

# write to csv
csvPath = "CSV\\"
df.to_csv(csvPath + keyword + date + '.csv', index=False)

stop_words = stopwords.words('english')
stop_words.extend(
    ['me', 'and', 'or', 'i', 'I', 'you', 'u', 'we', 'We', 'us', 'he', 'she', 'it', 'they', 'them', 'would', 'got',
     'went', 'could', 'made', 'give', 'go', 'get', 'also', 'much', 'must', 'well', 'much', 'nil', 'ok'])
from tabulate import tabulate


# Defining functions
def format_topics_sentences(ldamodel=None, corpus=None, texts=None, sourcedata=None):
    # Init output
    sent_topics_df = pd.DataFrame()

    # Get main topic in each document
    for i, row_list in enumerate(ldamodel[corpus]):
        row = row_list[0] if ldamodel.per_word_topics else row_list
        # print(row)
        row = sorted(row, key=lambda x: (x[1]), reverse=True)
        # Get the Dominant topic, Perc Contribution and Keywords for each document
        for j, (topic_num, prop_topic) in enumerate(row):
            if j == 0:  # => dominant topic
                wp = ldamodel.show_topic(topic_num)
                topic_keywords = ", ".join([word for word, prop in wp])
                sent_topics_df = sent_topics_df.append(
                    pd.Series([int(topic_num), round(prop_topic, 4), topic_keywords]), ignore_index=True)
            else:
                break
    sent_topics_df.columns = ['Dominant_Topic', 'Perc_Contribution', 'Topic_Keywords']

    # Add original text to the end of the output
    contents = pd.Series(sourcedata)
    sent_topics_df = pd.concat([sent_topics_df, contents], axis=1)
    return (sent_topics_df)


def my_preprocess(text):
    return simple_preprocess(text, deacc=True)  # deacc=True removes punctuations


def word_swap(text):
    # for item in text.split():
    if len(text.split()) < 5:
        text = " "
    else:
        text = text
    return (text)


'''
datetoday = input("Please enter today's date in 6 digits:\n")
print("")
filename = input("Please enter the name of the file without the extension:\n")
print("")
numtopics = input("Please enter the number of word clusters to generate:\n")
print("")
col_name = input("Please enter column name:\n")
'''

numberofquestions = 4
datetoday = datetime.today().strftime('%d%m%y')
filename = "covid19"
numtopics = 4
listOfQuestions = []
for x in range(numberofquestions):
    listOfQuestions.append("Question %d" %x)
print(listOfQuestions)
print(listOfQuestions[0])
OUTPUT_DIR = "output/"
current_directory = os.getcwd()
# directory = os.chdir(current_directory + "/rawdata")
directory = os.chdir(r"OBS_quality_quotes")
data = pd.read_excel(filename + ".xlsx", 'Master')

'''
"C01: 1.	What was the most challenging moment during the course? How did you overcome it?"	
"C02: 2.	How was your experience of going through the course with peers from a different school? What went well or did not go well?"	
"C03: 3.	What have you learnt about the environment / community?"	
"C04: 4.	What personal goals / resolutions do you have after leaving this course?"	
"D02.reason: 2.	If you could choose from the start, would you still choose to go through the MOE-OBS Challenge Programme?"	
"D03.reason: 3.	Do you think the MOE-OBS Challenge Course should be a shared experience for all Secondary 3 youths? "	
"F03: 3.	Do you have any other comments or suggestions for improvement? Be as specific as you can!"
'''

#columns_to_LDA = data[[col_name]]
#columns_to_LDA = data[['Q69_11','Q69_12','Q70_12','Q70_13','Q93_29','Q93_35','Q94_30','Q94_36']]
#columns_to_LDA = data[['(Q69) 12. Do you have suggestions for improvement for any of the Mission X activities?']]
columns_to_LDA = data[[questionNumber1, questionNumber2]]
# (Q69) 11. Do you have suggestions for improvement for any of the Mission X activities?    - BR1 - Q69_11
# (Q69) 12. Do you have suggestions for improvement for any of the Mission X activities?    - FV1 - Q69_12
# (Q70) 12. What other topics would you like to be included in the training workshop?       - BS1 - Q70_12
# (Q70) 13. What other topics would you like to be included in the training workshop?       - FW1 - Q70_13
# (Q93) 29. What were your key takeaways?                                                   - DA1 - Q93_29
# (Q93) 35. What were your key takeaways?                                                   - HJ1 - Q93_35
# (Q94) 30. Any other comments or areas of improvements for Mission X?                      - DB1 - Q94_30
# (Q94) 36. Any other comments or areas of improvements for Mission X?                      - HK1 - Q94_36

os.chdir(current_directory)

list_of_topictables = []
list_of_df_dominant_topic=[]
df = columns_to_LDA
i=0
j=0
# WORD THRESHOLD - Only analyse sentences which have more than 5 words
for column in columns_to_LDA.columns:
    df['source'] = columns_to_LDA[column].fillna('').astype(str).apply(word_swap)
    df['clean'] = df['source'].apply(my_preprocess)
    bigram = gensim.models.Phrases(df['clean'], min_count=2, threshold=10) # higher threshold fewer phrases.
    # Faster way to get a sentence clubbed as a trigram/bigram
    bigram_mod = gensim.models.phrases.Phraser(bigram)

    # STEP 4 REMOVE STOPWORDS, MAKE BIGRAMS AND LEMMATIZE -----------------------------
    def remove_stopwords(list_of_words):
        return [word for word in list_of_words if word not in stop_words]
    df['clean'] = df['clean'].apply(remove_stopwords)
    df['clean'] = df['clean'].apply(lambda list_of_words: bigram_mod[list_of_words])

    # STEP 11 CREATE THE DICTIONARY AND CORPUS NEEDED FOR TOPIC MODELLING -----------------------------
    # Create Dictionary
    id2word = corpora.Dictionary(df['clean'])

    # Create Corpus
    texts = df['clean']

    # Term Document Frequency
    corpus = [id2word.doc2bow(text) for text in texts]

    # STEP 5 BUILDING THE TOPIC MODEL -----------------------------
    # Build LDA model
    lda_model = gensim.models.ldamodel.LdaModel(corpus=corpus, id2word=id2word, num_topics=numtopics,random_state=100,
                                               update_every=1, chunksize=100, passes=50, alpha='auto',per_word_topics=True)
    #cm = CoherenceModel(model=lda_model, corpus=corpus, texts=texts, coherence='c_v', dictionary=id2word)
    #print(cm)

    top_words_per_topic = []
    for t in range(lda_model.num_topics):
        top_words_per_topic.extend([(t, ) + x for x in lda_model.show_topic(t, topn = 10)])
    topictable = pd.DataFrame(top_words_per_topic, columns=['Topic', 'Word', 'Weight',])
    questionNumbers = []
    for j in range(len(topictable.index)):
        questionNumbers.append(listOfQuestions[i])
    print(len(topictable.index))
    print(len(questionNumbers))
    topictable['Question'] = questionNumbers

    # topictable.columns = [column + '_' + str(col_header) for col_header in topictable.columns]
    list_of_topictables.append(topictable)
    print(list_of_topictables)
    # Visualize the topics using pyLDAvis
    vis = pyLDAvis.gensim.prepare(lda_model, corpus, id2word)

    # Save the visualisation into an interactive page. Details on how to read it here: https://www.machinelearningplus.com/nlp/topic-modeling-gensim-python/
    pyLDAvis.save_html(vis, f'{OUTPUT_DIR}filename_{column}_LDA_vis_{datetoday}.html')

    df_topic_sents_keywords = format_topics_sentences(ldamodel=lda_model,
                                                      corpus=corpus,
                                                      texts=df['clean'],
                                                      sourcedata=df['source']) #changed from df[column]

    # Format
    df_dominant_topic = df_topic_sents_keywords.reset_index()
    df_dominant_topic.columns = ['Document_No', 'Dominant_Topic', 'Topic_Perc_Contrib', 'Keywords', 'Text']
    # df_dominant_topic.columns = [column + '_' + str(col_header) for col_header in df_dominant_topic.columns]
    questionNumbers=[]
    for j in range(len(df_dominant_topic.index)):
        questionNumbers.append(listOfQuestions[i])
    df_dominant_topic['Question'] = questionNumbers
    list_of_df_dominant_topic.append(df_dominant_topic)
    i += 1

# write the topic tables and topic master matrix into one dataframe
alltopictables = pd.DataFrame()
alltopictables = pd.concat(list_of_topictables)

# N = 3
# alltopictables = np.split(alltopictables, np.arange(N, len(df.columns), N), axis=1)
#
# alltopictablesfinal = pd.DataFrame()
#
# alltopictablesfinal = pd.concat(alltopictables)
# print(alltopictablesfinal)
mastermatrix = pd.DataFrame()
mastermatrix = pd.concat(list_of_df_dominant_topic)
outputmatrix = pd.concat([data,mastermatrix])

# Writing the dataframes into a single Excel workbook ------------------------
outputfilename = f'{OUTPUT_DIR}filename{datetoday}_export.xlsx'
outputfilename_wo_ext = f'filename{datetoday}_export'
writer = pd.ExcelWriter(outputfilename, engine = 'xlsxwriter')
outputmatrix.to_excel(writer,sheet_name = "Masterlist")
mastermatrix.to_excel(writer,sheet_name = "Topic Master Matrix")
alltopictables.to_excel(writer,sheet_name = "Topic Tables")
writer.save()

print(f'All done')
