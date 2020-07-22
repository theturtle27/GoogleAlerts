import GetOldTweets3 as got
import pandas as pd
import openpyxl
from datetime import datetime
from googletrans import Translator

# specify path to output excel and csv
excelPath = r"Excel\\"
csvPath = r"CSV\\"

# get date time
now = datetime.now()
date = now.strftime("%d-%m-%y %H%M")

# open parameters workbook
wb_obj = openpyxl.load_workbook("Parameters.xlsx")

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# get keyword and count from sheet
keyword = sheet_obj.cell(row=10, column=2).value
count = int(sheet_obj.cell(row=11, column=2).value)

# Creation of translator object
translator = Translator()

# Creation of query object
tweetCriteria = got.manager.TweetCriteria().setQuerySearch(keyword) \
    .setMaxTweets(count)

# Creation of list that contains all tweets
tweets = got.manager.TweetManager.getTweets(tweetCriteria)

# Creating list of chosen tweet data
text_tweets = [[tweet.date, tweet.text, tweet.username, tweet.retweets] for tweet in tweets]

# Loop through tweets and change language to english
for i in range(len(text_tweets)):
    translation = translator.translate(text_tweets[i][1])
    text_tweets[i][1] = translation.text

# create dataFrame
df = pd.DataFrame(text_tweets, columns=["Time", "Text", "Username", "Retweets"])

# remove timezone info from Time column if not error
df['Time'] = df['Time'].dt.tz_localize(None)

# look for existing file
try:
    wb = openpyxl.load_workbook(excelPath + keyword + '.xlsx')

    # Inside this context manager, handle everything related to writing new data to the file\
    # without overwriting existing data
    with pd.ExcelWriter(excelPath + keyword + '.xlsx', engine='openpyxl') as writer:
        # Your loaded workbook is set as the "base of work"
        writer.book = wb

        # Loop through the existing worksheets in the workbook and map each title to\
        # the corresponding worksheet (that is, a dictionary where the keys are the\
        # existing worksheets' names and the values are the actual worksheets)
        writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}

        # Write the new data to the file without overwriting what already exists
        df.to_excel(writer, date, index=False)

        # Save the file
        writer.save()
# else write new file
except Exception as e:
    print(e)
    df.to_excel(excelPath + keyword + '.xlsx',
                sheet_name=date,
                index=False)

# write to csv
df.to_csv(csvPath + keyword + date + '.csv', index=False)
